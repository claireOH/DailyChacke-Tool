from datetime import timedelta, datetime
from fileinput import close

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from tkinter import *
from tkinter import ttk

from tkcalendar import *
import babel.numbers

import tkinter.messagebox as msbox

class report ():
    name        = ''
    date        = ''

    min_column  = 0
    max_column  = 0

    rangeStatus = False
    copyStatus  = False
    tabColor    = ''

    file        = None
    sheet       = None

    year        = 0
    mon         = 0
    startDay    = 0
    lastDay     = 0
    numOfDays   = 0

    fileName    = ''

    path        = ''
    
    checkReport = []
    warnReport  = []

    def __init__ (self):
        # title
        root = Tk()
        root.title("日報チェックツール")
        root.geometry("270x90")
        root.resizable(False, False)

        # Set the frame
        topFrame = Frame(root, relief="solid")
        topFrame.pack(side="top", fill="both")

        bottomFrame = Frame(root, relief="solid")
        bottomFrame.pack(side="bottom", fill="both")

        # input the schedule
        lbDate = Label(topFrame, text="チェック期間")
        lbDate.grid(row=1, column=0)
        entStDate = DateEntry(topFrame, width=11, foreground='white', borderwidth=4, 
                            locale='en_US', date_pattern='y/mm/dd')
        entStDate.config(state="readonly")
        entStDate.grid(row=1, column=1, sticky='ew')

        lbSeparator = Label(topFrame, text="~")
        lbSeparator.grid(row=1, column=2, columnspan=1)

        entLaDate = DateEntry(topFrame, width=11, foreground='white', borderwidth=4, 
                            locale='en_US', date_pattern='y/mm/dd')
        entLaDate.config(state="readonly")
        entLaDate.grid(row=1, column=3, sticky='ew')

        # Input the person
        # Select the number of people to check
        lbRange = Label(topFrame, text="コピー範囲")
        lbRange.grid(row=2, column=0)
        entRange = ttk.Combobox(topFrame, height=2, values=["入力データ不備", "入力データ不備 + 未入力"])
        entRange.set("入力データ不備")
        entRange.config(state="readonly")
        entRange.grid(row=2, column=1, columnspan=3, sticky='ew')

        # Input the file name
        lbFileName = Label(topFrame, text="結果ファイル名")
        lbFileName.grid(row=3, column=0)
        entFileName = Entry(topFrame, width=28)
        entFileName.grid(row=3, column=1, columnspan=3, sticky='ew')

        # set the button label
        lbBtn = Label(bottomFrame)
        lbBtn.pack()

        # set the check button
        btnCheck = Button(lbBtn, width=4, height=1, text="Check",
                        command=lambda: self.btnClickEvent_check(entStDate.get_date(), entLaDate.get_date(), 
                                                                entRange.get(), entFileName.get()))
        btnCheck.pack(side="left")

        # set the close button
        btnClose = Button(lbBtn, width=4, height=1, text="Close", command=root.destroy)
        btnClose.pack(side="left")

        # run the program
        root.mainloop()

    def btnClickEvent_check(self, argStDate, argLaDate, argRange, argFName):
        
            # 입력받은 데이터가 전부 유효한 경우
            if self.checkInputData(argStDate, argLaDate, argRange, argFName):
                # 시작 메시지 출력
                msbox.showinfo("開始", "日報チェックを開始します")

                # 일보 위치 경로 설정
                self.setPathAndMove()
                # print("경로설정 성공")

                # 확인할 일보 파일 추출
                self.sortReport()
                # print("파일추출 성공")

                # 결과 출력할 파일 생성
                self.setResult()
                # print("파일생성 성공")

                # 일보 내용 확인
                for report in self.checkReport:
                    self.checkReportData(report)
                    print(self.name + " : 完了")
                # print("전체확인 성공")

                # 2차 결과 파일 저장
                self.file.save(filename = self.fileName)
                # print("2차저장 성공")
                
                # 잘못된 파일 확인
                for warnReport in self.warnReport:
                    self.checkWarnReport(warnReport)
                # print("워닝확인 성공")

                # 최종 결과 파일 저장
                self.file.save(filename = self.fileName)
                # print("3차저장 성공")

                # 종료 메시지 출력
                msbox.showinfo("完了", "日報チェックを完了しました")

                # 변수 초기화
                self.valueReset()
                
    def valueReset (self):
        self.name        = ''
        self.date        = ''

        self.min_column  = 0
        self.max_column  = 0

        self.rangeStatus = False
        self.copyStatus  = False
        self.tabColor    = ''

        self.file        = None
        self.sheet       = None

        self.year        = 0
        self.mon         = 0
        self.startDay    = 0
        self.lastDay     = 0
        self.numOfDays   = 0

        self.fileName    = ''

        self.path        = ''
        
        self.checkReport = []
        self.warnReport  = []    
        
    def checkInputData (self, argStartDate, argLastDate, argRange, argFileName):
        try :
            # 첫날짜가 마지막날짜보다 늦을 경우 에러
            if (argStartDate <= argLastDate and 
                argStartDate.year == argLastDate.year and argStartDate.month == argLastDate.month):
                self.year       = argStartDate.year
                self.mon        = argStartDate.month
                self.startDay   = argStartDate.day
                self.lastDay    = argLastDate.day
                self.numOfDays  = argLastDate.day - argStartDate.day + 1
            else:
                self.msgErrorBox("選択した日付は正しくありません")
                return False
            
            # 확인할 범위 설정
            # 현재는 입력 데이터 값만 확인 가능
            if argRange == "入力データ不備 + 未入力":
                self.rangeStatus = True
            elif argRange == "入力データ不備" :
                self.rangeStatus = False

            # 파일명이 입력되었는지 확인
            # 파일명 미입력 시 에러
            if (argFileName == ''):
                self.msgErrorBox("ファイル名を入力してください")
                return False
            # 파일명에 사용할 수 없는 문자가 포함되어 있는지 확인
            # 이하의 문자가 포함되어 있을 경우 에러
            # / < > * ? : " | ￥
            elif(argFileName.find('<') != -1 or argFileName.find('>') != -1
                    or argFileName.find('*') != -1 or argFileName.find('/') != -1
                    or argFileName.find(':') != -1 or argFileName.find('"') != -1
                    or argFileName.find('|') != -1 or argFileName.find('￥') != -1):
                self.msgErrorBox('ファイル名には次の文字は使えません\n/ < > * ? : " | ￥')
                return False
            # 파일명이 유효할 경우 확장자 포함 파일명 설정
            else:
                self.fileName = argFileName + r'.xlsx'

            return True

        except :
            self.msgWarningBox()

    def setPathAndMove (self):
        # 파일 경로 설정
        # 1년의 기준은 4월
        # 1, 2, 3월은 전년도 폴더에 위치
        if self.mon < 4:
            setDateLink = '\%d年度\%s年%s月'%(self.year-1,self.year,self.mon)
        else:
            setDateLink = '\%s年度\%s年%s月'%(self.year,self.year,self.mon)

        link        = r"[Link]"
        self.path   = link + setDateLink

        # 설정한 경로로 이동
        try :
            os.chdir(self.path)
        except:
            self.msgWarningBox()

    def sortReport (self):
        # 설정된 파일 경로의 모든 파일 확인
        allFileList = os.listdir(os.getcwd())

        # 확인하려는 월을 세팅
        # 1~9월은 앞에 0을 붙인다. ex) 01, 02 ...
        if self.mon < 10:
            fileTitle = '%d0%d_'%(self.year, self.mon)
        else :
            fileTitle = '%d%d_'%(self.year, self.mon)

        # 확인해야하는 파일만 추출
        for file in allFileList:
            # 일보가 맞고, 확인하는 월이 포함되어 있을 경우
            if file.startswith("業務日報_") and fileTitle in file:
                self.checkReport.append(file)
            # 일보는 맞으나, 잘못된 파일명을 가진 일보일 경우
            elif file.startswith("業務日報_"):
                self.warnReport.append(file)
            # 일보가 아닌 경우는 패스
            else:
                continue

    def setResult (self):
        try :
            # 결과 파일을 생성
            self.file           = Workbook()
            self.sheet          = self.file.active
            self.sheet.title    = "チェック結果"

            # 소항목 타이틀 설정
            self.sheet.append(["入力者", "不足日", "勤務区分", "不足項目", "備考"])
        
            # 부족항목의 경우 너비를 다른 항목에 비해 넓게 설정
            self.sheet.column_dimensions["D"].width = 15
            self.sheet.column_dimensions["E"].width = 20

            # 소항목 셀 색상 설정 #FFFF00 
            self.sheet["A1"].fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
            self.sheet["B1"].fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
            self.sheet["C1"].fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
            self.sheet["D1"].fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
            self.sheet["E1"].fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")

            # 소항목 타이틀 폰트는 Bold설정
            self.sheet["A1"].font = Font(bold=True)
            self.sheet["B1"].font = Font(bold=True)
            self.sheet["C1"].font = Font(bold=True)
            self.sheet["D1"].font = Font(bold=True)
            self.sheet["E1"].font = Font(bold=True)

            # 결과 파일 1차 저장
            self.file.save(filename = self.fileName)

        # 결과 파일 생성에서 문제가 생겼을 경우 에러
        except :
            self.msgWarningBox()

    def checkReportData (self, argFile):
        # 일보 열기
        reportFile     = load_workbook(filename = argFile, data_only=True)
        reportSheet    = reportFile['日報入力']

        # 이름, 날짜 확인
        self.name = reportSheet['C2'].value
        self.date = reportSheet['C3'].value.strftime('%Y/%m')

        # 확인할 일 세팅
        self.min_column = self.startDay + 2
        self.max_column = self.min_column + self.numOfDays

        # 복사상태 리셋
        self.copyStatus = False

        for col in range(self.min_column, self.max_column, 1):
            # row16 : 날짜, row17 : 시프트, row20 : 입력데이터
            row16 = reportSheet.cell(16, col)
            row19 = reportSheet.cell(19, col)
            row20 = reportSheet.cell(20, col)

            # 현재 확인하는 날짜
            dateCheckingNow = row16.value.strftime('%m/%d')

            # 휴일은 패스
            if (row19.value == '法休' or row19.value == '休' 
                or row19.value == '有' or row19.value == '明'):
                continue
            # 미입력일 경우, 또는 일부 미입력일 경우
            elif row20.value == '-':
                self.checkMemoOrComment(dateCheckingNow, row16, row19, row20, "未入力")
            # 입력된 데이터가 올바른지 확인
            else:
                # row26 : 근무시간 정확도, row27 : 근무형태 정확도
                row26 = reportSheet.cell(26, col).value
                row27 = reportSheet.cell(27, col).value

                # 근무시간 정확도 확인
                if row26 != 'None' and row26 != '#N/A' and row26 != '◯' and row26 != '×':
                    self.sheet.append([self.name, dateCheckingNow, row19.value, "時間確認要", "-"])
                    self.copyStatus = True
                    self.tabColor   = 'R'
                elif row26 == '#N/A' or row26 == '' or row26 == None:
                    continue
                elif row26 == '×':
                    # 메모나 코멘트가 있을 경우
                    if (row16.comment or row19.comment or row20.comment):
                        self.checkMemoOrComment(dateCheckingNow, row16, row19, row20, "時間不一致")
                    else:
                        self.sheet.append([self.name, dateCheckingNow, row19.value, "時間不一致", "-"])
                    self.copyStatus = True
                    self.tabColor   = 'R'

                # 근무형태 정확도 확인
                if row27 != 'None' and row27 != '#N/A' and row27 != '◯' and row27 != '〇' and row27 != '×':
                    self.sheet.append([self.name, dateCheckingNow, row19.value, "勤務区分確認要", "-"])
                    self.copyStatus = True
                    self.tabColor   = 'R'
                elif row27 == '#N/A'or row27 == '' or row27 == None:
                    continue
                elif row27 == '×':
                    if (row16.comment or row19.comment or row20.comment):
                        self.checkMemoOrComment(dateCheckingNow, row16, row19, row20, "勤務区分不一致")
                    else:
                        self.sheet.append([self.name, dateCheckingNow, row19.value, "勤務区分不一致", "-"])
                    self.copyStatus = True
                    self.tabColor   = 'R'
        
        # 입력사항이 있을 경우 일보 시트 복사
        if self.copyStatus:
            id_sheet = self.file.create_sheet(self.name)
            
            # 일보 시트 탭 색상 설정
            # 입력내용 불일치, 상세확인요 경우 Red
            if self.tabColor == 'R':
                id_sheet.sheet_properties.tabColor = "FF0000"
            # 미입력의 경우 Yellow
            elif self.tabColor == 'Y':
                id_sheet.sheet_properties.tabColor = "FFFF00"

            # 셀 색상 설정
            fill = PatternFill("solid", start_color="FF0000", end_color="FF0000")
            
            for copy_row in range(1, reportSheet.max_row + 1):
                for copy_column in range(1, reportSheet.max_column + 1):
                    tmp = reportSheet.cell(row=copy_row, column=copy_column).value

                    # 입력월
                    if copy_row == 3 and copy_column == 3:
                        id_sheet.cell(row=copy_row, column=copy_column).value = tmp.strftime('%Y/%m/%d')
                    # 날짜
                    elif copy_row == 16 and copy_column > 2 and tmp != None:
                        id_sheet.cell(row=copy_row, column=copy_column).value = tmp.strftime('%m/%d')
                    # 시간
                    elif copy_row > 20 and copy_row < 25:
                        if copy_column > 2 and tmp != None:
                            set_tmp =  ':'.join((str(tmp)).split(':')[:2])
                            id_sheet.cell(row=copy_row, column=copy_column).value = set_tmp
                    # 그 외
                    else :
                        id_sheet.cell(row=copy_row, column=copy_column).value = tmp

                    # 셀 서식 설정
                    # A열 너비
                    if copy_column == 1:
                        id_sheet.column_dimensions["A"].width = 2.5
                    # 셀 높이
                    if copy_row == 5:
                        id_sheet.row_dimensions[copy_row].height = 0.1
                    # 22행, 23행 
                    elif copy_row == 22 or copy_row == 23:
                        id_sheet.row_dimensions[copy_row].height = 0.1
                    # 셀 컬러
                    if copy_row == 26 or copy_row == 27:
                        if tmp == '×':
                            id_sheet.cell(row=copy_row, column=copy_column).fill = fill


        # 파일 종료
        reportFile.close()
    
    def checkMemoOrComment (self, argdateCheckingNow, argR16, argR19, argR20, argMsg):
        text = ''

        # 메모나 코멘트가 붙어 있는 지 확인
        if argR16.comment or argR19.comment or argR20.comment:
            if argR16.comment:
                text = argR16.comment.text
            elif argR19.comment:
                text = argR19.comment.text
            else:
                text = argR20.comment.text

            # 코멘트일 경우, 시스템 안내 문구 제외하고 추출
            # 시스템 안내 문구는 172자
            if(text != '' and len(text) > 172):
                memo = text[172:len(text)]
                self.sheet.append([self.name, argdateCheckingNow, argR19.value, argMsg, memo])
            # 메모일 경우 작성자 이름 제외하고 추출
            # 구분자 :
            elif text != '' :
                try :
                    writer, memo = text.split(':', 2)
                    # 관리자가 작성한 메모는 제외
                    administrator1 = '[Name]'
                    administrator2 = '[Name]'

                    if (writer != administrator1 or writer != administrator2):
                        self.sheet.append([self.name, argdateCheckingNow, argR19.value, argMsg, memo])
                except :
                        self.sheet.append([self.name, argdateCheckingNow, argR19.value, argMsg, text])
        else :
            self.sheet.append([self.name, argdateCheckingNow, argR19.value, argMsg, "-"])
            # 미입력의 경우 처음 복사 설정 여부에 따라 결정
            if self.rangeStatus:
                self.copyStatus = True
                self.tabColor   = 'Y'
            else :
                self.copyStatus = False
                self.tabColor   = ''

    def checkWarnReport (self, argFile):
        id_warnFile = load_workbook(filename = argFile, data_only=True)
        id_warnFileSheet = id_warnFile['日報入力']

        warnName = id_warnFileSheet['C2'].value

        id_warnFileSheet.append([warnName, "-", "-", "-", "ファイル名修正要"])

    def msgErrorBox(self, argMsg):
        msbox.showerror("Error", argMsg)

    def msgWarningBox (self):
        msbox.showwarning("Warning", "エラーが発生しました")


report()
